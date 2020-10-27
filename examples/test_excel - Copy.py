"""from openpyxl import load_workbook
workbook = load_workbook(filename="data\CRQ_ECPG_Element_list.xls")
workbook.sheetnames
sheet = workbook.active
"""
base_file = 'data\CRQ_ECPG_Element_list.xls'
dest_location = 'C:\\Users\\srinivasan.c\\Desktop'
data_location = 'C:\\Users\\srinivasan.c\\Desktop\\ECPG_list.txt'
"""
07/01/2020	12:54:22	AM	55812	NewCardRequest.cfm	c:\ecpg\BL_Publish_Staging_CFCS_CCRT_16\
['c:\\ecpg\\BL_Publish_Staging_CFCS_CCRT_16\\NewCardRequesasdfsdft.cfm', '20200701',
 '12:54:22', '55812']
"""
import shutil
import os, os.path
import re
import pandas as pd
#import win32com.client
def split_data(data_array):
    date_ = data_array[1][4:6] + '/' + data_array[1][6:8] + '/' + data_array[1][:4]
    time_ = data_array[2]
    size_ = data_array[3]
    last_slash = data_array[0].rfind('\\')
    path_ = data_array[0][:last_slash+1]
    file_ = data_array[0][last_slash+1:]
    sect_ = 'AM'
    return date_, time_ , sect_, size_, path_, file_



data_read = open(data_location, "r")
got_error = False
CRQ_no = ''
dataframe_array = pd.DataFrame(columns=['date', 'time', 'sect', 'size', 'path', 'file'])
for line_read in data_read:
    line_read = line_read.strip()
    line_read = re.sub(' +', ' ',line_read.rstrip('\n'))
    if line_read.strip() != '':
        data_split = line_read.split(' ')
        print(data_split,len(data_split),line_read.rfind('\\'))
        if len(data_split) == 4 and line_read.rfind('\\') != -1:
            date_, time_ , sect_, size_, path_, file_ = split_data(data_split)
            dataframe_array = dataframe_array.append({'date':date_, 'time':time_, 'sect':sect_, 'size':size_, 'file':file_, 'path':path_}, ignore_index=True)
        else:
            if len(data_split) == 1 and data_split[0].isnumeric():
                CRQ_no = str(data_split[0])
            else:
                got_error = True
                print("Correct the input data")
                break
if got_error != True :
    if CRQ_no == '':
        CRQ_no = input("Enter CRQ Number:")
    new_file_name = dest_location + '\\CRQ_' + CRQ_no + '_ECPG_Element_list.xls'
    newPath = shutil.copy(base_file, dest_location)
    os.rename(newPath, new_file_name)
    
    import xlwings as xw
    wb = xw.Book(new_file_name)
    sht = wb.sheets['Element List']
    y = dataframe_array.iloc[:,:].values
    sht.range('A3').value = y
    #wb.save()
    #wb.macro('Validate')
    #wb.set_mock_caller()
    #print(dataframe_array.iloc[:,:].values )
